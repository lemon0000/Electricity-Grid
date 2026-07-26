"""Validate the U.S. major power outage event dataset and source boundaries."""

from __future__ import annotations

import argparse
import hashlib
import io
import json
import zipfile
from collections import Counter
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree

import openpyxl
import yaml


XLINK_HREF = "{http://www.w3.org/1999/xlink}href"


def _verify_sha256_manifest(root: Path) -> bool:
    manifest_path = root / "SHA256SUMS"
    if not manifest_path.exists():
        return False
    expected_paths = set()
    for line in manifest_path.read_text(encoding="ascii").splitlines():
        expected, relative = line.split("  ", maxsplit=1)
        expected_paths.add(relative)
        path = root / Path(relative)
        if not path.is_file() or hashlib.sha256(path.read_bytes()).hexdigest() != expected:
            return False
    actual_paths = {
        path.relative_to(root).as_posix()
        for path in root.rglob("*")
        if path.is_file()
        and path.name != "SHA256SUMS"
        and not path.name.endswith(".partial")
    }
    return expected_paths == actual_paths


def _local_name(element: ElementTree.Element) -> str:
    return element.tag.rsplit("}", maxsplit=1)[-1]


def _article_metadata(article_path: Path) -> dict[str, object]:
    root = ElementTree.parse(article_path).getroot()
    article_ids = {
        element.attrib.get("pub-id-type"): element.text
        for element in root.iter()
        if _local_name(element) == "article-id"
    }
    licenses = [
        element for element in root.iter() if _local_name(element) == "license"
    ]
    media = [
        element.attrib.get(XLINK_HREF)
        for element in root.iter()
        if _local_name(element) == "media"
    ]
    return {
        "article_ids": article_ids,
        "license_type": licenses[0].attrib.get("license-type") if licenses else None,
        "license_url": licenses[0].attrib.get(XLINK_HREF) if licenses else None,
        "supplementary_members": sorted(value for value in media if value),
    }


def _is_missing(value: object) -> bool:
    return value is None or (
        isinstance(value, str) and value.strip().upper() == "NA"
    )


def _combine(date_value: datetime, time_value: object) -> datetime:
    return datetime.combine(date_value.date(), time_value)


def run(config_path: Path, *, write_output: bool = True) -> dict[str, object]:
    config = yaml.safe_load(config_path.read_text(encoding="utf-8"))
    root = Path(config["source"]["path"])
    expected = config["expected"]
    supplement_path = root / "supplementary_files.zip"
    supplement_bytes = supplement_path.read_bytes()
    supplement_container_sha256 = hashlib.sha256(supplement_bytes).hexdigest()

    article = _article_metadata(root / "article.xml")
    with zipfile.ZipFile(io.BytesIO(supplement_bytes)) as archive:
        members = sorted(archive.namelist())
        member_summary = {
            member: {
                "bytes": len(content := archive.read(member)),
                "sha256": hashlib.sha256(content).hexdigest(),
            }
            for member in members
        }
        workbook_bytes = archive.read(expected["workbook_member"])
    workbook = openpyxl.load_workbook(
        io.BytesIO(workbook_bytes),
        read_only=True,
        data_only=True,
    )
    worksheet = workbook[expected["sheet"]]
    headers = [cell.value for cell in worksheet[expected["header_row"]]]
    if len(headers) != len(set(headers)):
        raise ValueError("Outage workbook headers must be unique")
    missing_columns = set(expected["required_columns"]) - set(headers)
    if missing_columns:
        raise ValueError(f"Missing outage columns: {sorted(missing_columns)}")
    records = [
        dict(zip(headers, row))
        for row in worksheet.iter_rows(
            min_row=expected["data_start_row"],
            values_only=True,
        )
    ]

    observation_ids = [record["OBS"] for record in records]
    years = [record["YEAR"] for record in records]
    postal_jurisdictions = sorted({record["POSTAL.CODE"] for record in records})
    missing_counts = {
        column: sum(_is_missing(record[column]) for record in records)
        for column in expected["missing"]
    }
    cause_counts = Counter(record["CAUSE.CATEGORY"] for record in records)
    candidate_event_counts = Counter(
        tuple(record[column] for column in expected["candidate_event_key"])
        for record in records
    )
    duplicate_candidate_counts = [
        count for count in candidate_event_counts.values() if count > 1
    ]
    duplicate_candidate_event_groups = len(duplicate_candidate_counts)
    duplicate_candidate_rows = sum(duplicate_candidate_counts)
    duplicate_candidate_excess_rows = sum(
        count - 1 for count in duplicate_candidate_counts
    )

    known_demand_losses = [
        float(record["DEMAND.LOSS.MW"])
        for record in records
        if not _is_missing(record["DEMAND.LOSS.MW"])
    ]
    positive_demand_loss_records = sum(value > 0 for value in known_demand_losses)
    zero_demand_loss_records = sum(value == 0 for value in known_demand_losses)
    negative_demand_loss_records = sum(value < 0 for value in known_demand_losses)

    temporal_columns = (
        "OUTAGE.START.DATE",
        "OUTAGE.START.TIME",
        "OUTAGE.RESTORATION.DATE",
        "OUTAGE.RESTORATION.TIME",
        "OUTAGE.DURATION",
    )
    duration_differences = []
    complete_temporal_records = 0
    zero_duration_records = 0
    zero_duration_timestamp_consistent_records = 0
    zero_duration_cause_counts: Counter[object] = Counter()
    negative_duration_records = 0
    negative_timestamp_duration_records = 0
    for record in records:
        duration = record["OUTAGE.DURATION"]
        if not _is_missing(duration):
            zero_duration_records += float(duration) == 0.0
            if float(duration) == 0.0:
                zero_duration_cause_counts[record["CAUSE.CATEGORY"]] += 1
            negative_duration_records += float(duration) < 0.0
        if any(_is_missing(record[column]) for column in temporal_columns):
            continue
        complete_temporal_records += 1
        start = _combine(record["OUTAGE.START.DATE"], record["OUTAGE.START.TIME"])
        restoration = _combine(
            record["OUTAGE.RESTORATION.DATE"],
            record["OUTAGE.RESTORATION.TIME"],
        )
        timestamp_duration = (restoration - start).total_seconds() / 60.0
        negative_timestamp_duration_records += timestamp_duration < 0.0
        if float(duration) == 0.0 and abs(timestamp_duration) <= 1e-9:
            zero_duration_timestamp_consistent_records += 1
        difference = float(duration) - timestamp_duration
        if abs(difference) > 1e-9:
            duration_differences.append(int(round(difference)))
    duration_mismatch_counts = {
        str(key): value for key, value in sorted(Counter(duration_differences).items())
    }

    checks = {
        "sha256_manifest": _verify_sha256_manifest(root),
        "supplement_members": members == sorted(expected["supplement_members"])
        and member_summary == expected["supplement_members"],
        "article_doi": article["article_ids"].get("doi") == config["source"]["doi"],
        "article_pmcid": article["article_ids"].get("pmcid")
        == config["source"]["pmcid"],
        "article_license": article["license_type"] == "CC BY"
        and article["license_url"] == config["source"]["license_url"],
        "article_supplements": article["supplementary_members"]
        == sorted(expected["supplement_members"]),
        "worksheet_shape": worksheet.max_row == expected["worksheet_rows"]
        and worksheet.max_column == expected["worksheet_columns"],
        "source_rows": len(records) == expected["source_rows"],
        "observation_ids": len(set(observation_ids)) == len(observation_ids)
        and min(observation_ids) == expected["first_observation"]
        and max(observation_ids) == expected["last_observation"],
        "year_range": min(years) == expected["first_year"]
        and max(years) == expected["last_year"],
        "postal_jurisdictions": postal_jurisdictions
        == expected["postal_jurisdictions"],
        "nerc_region_count": len({record["NERC.REGION"] for record in records})
        == expected["nerc_regions"],
        "duplicate_candidate_events": duplicate_candidate_event_groups
        == expected["duplicate_candidate_event_groups"]
        and duplicate_candidate_rows == expected["duplicate_candidate_rows"]
        and duplicate_candidate_excess_rows
        == expected["duplicate_candidate_excess_rows"],
        "missingness": missing_counts == expected["missing"],
        "cause_counts": dict(cause_counts) == expected["cause_counts"],
        "complete_temporal_records": complete_temporal_records
        == expected["complete_temporal_records"],
        "duration_mismatch_minutes": duration_mismatch_counts
        == expected["duration_mismatch_minutes"],
        "zero_duration_records": zero_duration_records
        == expected["zero_duration_records"],
        "zero_duration_timestamp_consistency": (
            zero_duration_timestamp_consistent_records
            == expected["zero_duration_timestamp_consistent_records"]
            == zero_duration_records
        ),
        "zero_duration_cause_counts": dict(zero_duration_cause_counts)
        == expected["zero_duration_cause_counts"],
        "demand_loss_counts": len(known_demand_losses)
        == expected["known_demand_loss_records"]
        and positive_demand_loss_records
        == expected["positive_demand_loss_records"]
        and zero_demand_loss_records == expected["zero_demand_loss_records"]
        and negative_demand_loss_records
        == expected["negative_demand_loss_records"],
        "no_negative_duration": negative_duration_records == 0,
        "no_negative_timestamp_duration": (
            negative_timestamp_duration_records
            == expected["negative_timestamp_duration_records"]
            == 0
        ),
    }
    if not all(checks.values()):
        raise RuntimeError(f"U.S. outage data validation failed: {checks}")

    summary = {
        "dataset": config["source"]["dataset"],
        "pmcid": config["source"]["pmcid"],
        "doi": config["source"]["doi"],
        "license": config["source"]["license"],
        "checks": checks,
        "supplement_container_bytes": len(supplement_bytes),
        "supplement_container_sha256": supplement_container_sha256,
        "supplement_container_byte_stable": expected[
            "supplement_container_byte_stable"
        ],
        "supplement_members": member_summary,
        "source_rows": len(records),
        "first_year": min(years),
        "last_year": max(years),
        "postal_jurisdictions": postal_jurisdictions,
        "postal_jurisdiction_count": len(postal_jurisdictions),
        "nerc_regions": len({record["NERC.REGION"] for record in records}),
        "candidate_event_key": expected["candidate_event_key"],
        "duplicate_candidate_event_groups": duplicate_candidate_event_groups,
        "duplicate_candidate_rows": duplicate_candidate_rows,
        "duplicate_candidate_excess_rows": duplicate_candidate_excess_rows,
        "missing": missing_counts,
        "cause_counts": dict(cause_counts),
        "complete_temporal_records": complete_temporal_records,
        "duration_mismatch_minutes": duration_mismatch_counts,
        "zero_duration_records": zero_duration_records,
        "zero_duration_timestamp_consistent_records": (
            zero_duration_timestamp_consistent_records
        ),
        "zero_duration_cause_counts": dict(zero_duration_cause_counts),
        "negative_duration_records": negative_duration_records,
        "negative_timestamp_duration_records": negative_timestamp_duration_records,
        "known_demand_loss_records": len(known_demand_losses),
        "positive_demand_loss_records": positive_demand_loss_records,
        "zero_demand_loss_records": zero_demand_loss_records,
        "negative_demand_loss_records": negative_demand_loss_records,
        "evidence": config["evidence"],
    }
    if write_output:
        output_path = Path(config["output"]["summary_path"])
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(
            json.dumps(summary, indent=2, sort_keys=True) + "\n",
            encoding="utf-8",
        )
    return summary


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--config",
        type=Path,
        default=Path("configs/us_major_power_outages.yaml"),
    )
    args = parser.parse_args()
    print(json.dumps(run(args.config), indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
